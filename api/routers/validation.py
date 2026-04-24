"""
Validation & QC router
"""

import io
import tempfile
from typing import List

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from fastapi import APIRouter, File, Form, UploadFile, HTTPException
from fastapi.responses import StreamingResponse

from api.utils import save_uploads
from powertech_tools.utils.file_parser import read_headers_only, load_maxmin_for_plot
from powertech_tools.data.validator import validate_maxmin_file

router = APIRouter()

_GREEN = PatternFill("solid", fgColor="C6EFCE")
_RED   = PatternFill("solid", fgColor="FFC7CE")


@router.post("/headers")
async def get_headers(files: List[UploadFile] = File(...)):
    tmpdir = tempfile.mkdtemp(prefix="jerry_val_")
    try:
        paths = await save_uploads(files, tmpdir)
        headers, _, _, _ = read_headers_only(paths[0])
        return {"headers": headers}
    except Exception as e:
        raise HTTPException(500, str(e))


@router.post("/validate")
async def validate(
    files: List[UploadFile] = File(...),
    cycle_col: str = Form(...),
    limits_json: str = Form(...),
    # limits_json: {"VarName": {"min_lower": f, "min_upper": f, "max_lower": f, "max_upper": f}}
):
    """Validate max/min file against limits; return Excel with colour coding."""
    import json

    tmpdir = tempfile.mkdtemp(prefix="jerry_val_")
    try:
        paths = await save_uploads(files, tmpdir)
        limits = json.loads(limits_json)

        df, int_cols, int_to_disp, _ = load_maxmin_for_plot(paths[0])
        # Rebuild display-name df
        disp_df = df.copy()
        disp_df.columns = [int_to_disp.get(c, c) for c in disp_df.columns]

        results_df, summary = validate_maxmin_file(disp_df, limits, cycle_col)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            results_df.to_excel(writer, sheet_name="Results", index=False)
            disp_df.to_excel(writer, sheet_name="Data", index=False)
            pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)

            ws = writer.sheets["Results"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                status_cell = row[1]  # Status column
                fill = _GREEN if status_cell.value == "PASS" else _RED
                for cell in row:
                    cell.fill = fill

        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="validation_results.xlsx"',
                "X-Total":  str(summary["total_cycles"]),
                "X-Passed": str(summary["passed_cycles"]),
                "X-Failed": str(summary["failed_cycles"]),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
